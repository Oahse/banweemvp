from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_, desc, asc, text, case
from sqlalchemy.orm import aliased, selectinload
from typing import List, Tuple, Dict, Any, Optional
from datetime import datetime, timedelta, date
from decimal import Decimal
import csv
import io
import json
from dataclasses import dataclass

# Import models
from models.user import User
from models.product import Product, ProductVariant, Category
from models.order import Order, OrderItem
from models.review import Review
from models.subscription import Subscription
from models.transaction import Transaction
from models.payment_intent import PaymentIntent
from models.analytics import SubscriptionAnalytics, PaymentAnalytics


@dataclass
class DateRange:
    """Date range for analytics queries"""
    start_date: datetime
    end_date: datetime


@dataclass
class AnalyticsFilters:
    """Filters for analytics queries"""
    user_segment: Optional[str] = None
    subscription_status: Optional[str] = None
    payment_method: Optional[str] = None
    currency: Optional[str] = None
    country: Optional[str] = None
    category: Optional[str] = None


@dataclass
class SubscriptionMetrics:
    """Subscription metrics data structure"""
    total_active_subscriptions: int
    new_subscriptions: int
    canceled_subscriptions: int
    paused_subscriptions: int
    total_revenue: Decimal
    average_subscription_value: Decimal
    monthly_recurring_revenue: Decimal
    churn_rate: float
    conversion_rate: float
    retention_rate: float
    currency: str
    plan_breakdown: Dict[str, Any]
    geographic_breakdown: Dict[str, Any]


@dataclass
class CLVAnalysis:
    """Customer Lifetime Value analysis"""
    average_clv: Decimal
    median_clv: Decimal
    clv_by_segment: Dict[str, Decimal]
    clv_trend: List[Dict[str, Any]]
    top_customers: List[Dict[str, Any]]


@dataclass
class PaymentAnalytics:
    """Payment analytics data structure"""
    total_payments: int
    successful_payments: int
    failed_payments: int
    success_rate: float
    total_volume: Decimal
    average_payment_amount: Decimal
    breakdown_by_method: Dict[str, Any]
    breakdown_by_country: Dict[str, Any]
    failure_analysis: Dict[str, Any]


@dataclass
class RevenueForecast:
    """Revenue forecast data structure"""
    forecast_data: List[Dict[str, Any]]
    confidence_intervals: Dict[str, Any]
    growth_rate: float
    seasonal_factors: Dict[str, float]


class AnalyticsService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def get_subscription_metrics(
        self, 
        date_range: DateRange, 
        filters: Optional[AnalyticsFilters] = None
    ) -> SubscriptionMetrics:
        """
        Get comprehensive subscription metrics with real data analysis.
        
        Requirements: 12.1, 12.2, 12.3, 12.4, 12.5
        """
        if filters is None:
            filters = AnalyticsFilters()
        
        # Base date conditions
        date_conditions = and_(
            Subscription.created_at >= date_range.start_date,
            Subscription.created_at <= date_range.end_date
        )
        
        # Build filter conditions
        filter_conditions = []
        if filters.subscription_status:
            filter_conditions.append(Subscription.status == filters.subscription_status)
        if filters.currency:
            filter_conditions.append(Subscription.currency == filters.currency)
        
        # Combine all conditions
        all_conditions = [date_conditions] + filter_conditions
        where_clause = and_(*all_conditions) if all_conditions else date_conditions
        
        # Total active subscriptions
        active_query = select(func.count(Subscription.id)).where(
            and_(Subscription.status == "active", *filter_conditions) if filter_conditions 
            else Subscription.status == "active"
        )
        active_result = await self.db.execute(active_query)
        total_active = active_result.scalar() or 0
        
        # New subscriptions in date range
        new_query = select(func.count(Subscription.id)).where(where_clause)
        new_result = await self.db.execute(new_query)
        new_subscriptions = new_result.scalar() or 0
        
        # Canceled subscriptions in date range
        canceled_conditions = and_(
            Subscription.cancelled_at >= date_range.start_date,
            Subscription.cancelled_at <= date_range.end_date,
            *filter_conditions
        ) if filter_conditions else and_(
            Subscription.cancelled_at >= date_range.start_date,
            Subscription.cancelled_at <= date_range.end_date
        )
        
        canceled_query = select(func.count(Subscription.id)).where(canceled_conditions)
        canceled_result = await self.db.execute(canceled_query)
        canceled_subscriptions = canceled_result.scalar() or 0
        
        # Paused subscriptions
        paused_query = select(func.count(Subscription.id)).where(
            and_(Subscription.status == "paused", *filter_conditions) if filter_conditions
            else Subscription.status == "paused"
        )
        paused_result = await self.db.execute(paused_query)
        paused_subscriptions = paused_result.scalar() or 0
        
        # Revenue calculations
        revenue_query = select(func.sum(Subscription.price)).where(
            and_(Subscription.status.in_(["active", "paused"]), *filter_conditions) if filter_conditions
            else Subscription.status.in_(["active", "paused"])
        )
        revenue_result = await self.db.execute(revenue_query)
        total_revenue = Decimal(str(revenue_result.scalar() or 0))
        
        # Average subscription value
        avg_query = select(func.avg(Subscription.price)).where(
            and_(Subscription.status.in_(["active", "paused"]), *filter_conditions) if filter_conditions
            else Subscription.status.in_(["active", "paused"])
        )
        avg_result = await self.db.execute(avg_query)
        avg_value = Decimal(str(avg_result.scalar() or 0))
        
        # Monthly recurring revenue (active subscriptions only)
        mrr_query = select(func.sum(
            case(
                (Subscription.billing_cycle == "monthly", Subscription.price),
                (Subscription.billing_cycle == "yearly", Subscription.price / 12),
                else_=Subscription.price
            )
        )).where(
            and_(Subscription.status == "active", *filter_conditions) if filter_conditions
            else Subscription.status == "active"
        )
        mrr_result = await self.db.execute(mrr_query)
        mrr = Decimal(str(mrr_result.scalar() or 0))
        
        # Calculate churn rate (canceled / (active + canceled))
        total_for_churn = total_active + canceled_subscriptions
        churn_rate = (canceled_subscriptions / total_for_churn * 100) if total_for_churn > 0 else 0.0
        
        # Calculate conversion rate (need to get total users who viewed subscriptions)
        # For now, use a simplified calculation: new subscriptions / total users in period
        user_query = select(func.count(User.id)).where(
            and_(
                User.created_at >= date_range.start_date,
                User.created_at <= date_range.end_date
            )
        )
        user_result = await self.db.execute(user_query)
        total_users = user_result.scalar() or 0
        conversion_rate = (new_subscriptions / total_users * 100) if total_users > 0 else 0.0
        
        # Calculate retention rate (active / (active + canceled))
        retention_rate = (total_active / total_for_churn * 100) if total_for_churn > 0 else 0.0
        
        # Plan breakdown
        plan_query = select(
            Subscription.plan_id,
            func.count(Subscription.id).label("count"),
            func.sum(Subscription.price).label("revenue")
        ).where(
            and_(Subscription.status.in_(["active", "paused"]), *filter_conditions) if filter_conditions
            else Subscription.status.in_(["active", "paused"])
        ).group_by(Subscription.plan_id)
        
        plan_result = await self.db.execute(plan_query)
        plan_breakdown = {
            row.plan_id: {
                "count": row.count,
                "revenue": float(row.revenue or 0)
            }
            for row in plan_result.all()
        }
        
        # Geographic breakdown (using user country)
        geo_query = select(
            User.country,
            func.count(Subscription.id).label("count"),
            func.sum(Subscription.price).label("revenue")
        ).join(User, Subscription.user_id == User.id).where(
            and_(Subscription.status.in_(["active", "paused"]), *filter_conditions) if filter_conditions
            else Subscription.status.in_(["active", "paused"])
        ).group_by(User.country)
        
        geo_result = await self.db.execute(geo_query)
        geographic_breakdown = {
            row.country or "Unknown": {
                "count": row.count,
                "revenue": float(row.revenue or 0)
            }
            for row in geo_result.all()
        }
        
        # Get currency from first active subscription or default to USD
        currency_query = select(Subscription.currency).where(
            Subscription.status == "active"
        ).limit(1)
        currency_result = await self.db.execute(currency_query)
        currency = currency_result.scalar() or "USD"
        
        return SubscriptionMetrics(
            total_active_subscriptions=total_active,
            new_subscriptions=new_subscriptions,
            canceled_subscriptions=canceled_subscriptions,
            paused_subscriptions=paused_subscriptions,
            total_revenue=total_revenue,
            average_subscription_value=avg_value,
            monthly_recurring_revenue=mrr,
            churn_rate=churn_rate,
            conversion_rate=conversion_rate,
            retention_rate=retention_rate,
            currency=currency,
            plan_breakdown=plan_breakdown,
            geographic_breakdown=geographic_breakdown
        )

    async def calculate_customer_lifetime_value(
        self, 
        user_id: Optional[str] = None, 
        segment: Optional[str] = None
    ) -> CLVAnalysis:
        """
        Calculate customer lifetime value using actual transaction data.
        
        Requirements: 12.2, 12.3
        """
        # Base query for CLV calculation
        # CLV = (Average Order Value × Purchase Frequency × Gross Margin × Lifespan)
        # Simplified: Total revenue per customer / customer lifespan in months
        
        base_query = select(
            User.id,
            User.created_at,
            func.sum(Transaction.amount).label("total_spent"),
            func.count(Transaction.id).label("transaction_count"),
            func.max(Transaction.created_at).label("last_transaction"),
            func.min(Transaction.created_at).label("first_transaction")
        ).join(
            Transaction, User.id == Transaction.user_id
        ).where(
            Transaction.status == "succeeded"
        )
        
        # Apply filters
        if user_id:
            base_query = base_query.where(User.id == user_id)
        
        if segment:
            if segment == "high_value":
                base_query = base_query.having(func.sum(Transaction.amount) > 1000)
            elif segment == "frequent":
                base_query = base_query.having(func.count(Transaction.id) > 5)
            elif segment == "recent":
                recent_date = datetime.now() - timedelta(days=90)
                base_query = base_query.having(func.max(Transaction.created_at) > recent_date)
        
        base_query = base_query.group_by(User.id, User.created_at)
        
        result = await self.db.execute(base_query)
        customer_data = result.all()
        
        if not customer_data:
            return CLVAnalysis(
                average_clv=Decimal("0"),
                median_clv=Decimal("0"),
                clv_by_segment={},
                clv_trend=[],
                top_customers=[]
            )
        
        # Calculate CLV for each customer
        clv_values = []
        customer_clvs = []
        
        for row in customer_data:
            # Calculate customer lifespan in months
            if row.last_transaction and row.first_transaction:
                lifespan_days = (row.last_transaction - row.first_transaction).days
                lifespan_months = max(lifespan_days / 30.0, 1.0)  # Minimum 1 month
            else:
                lifespan_months = 1.0
            
            # Simple CLV calculation: total spent / lifespan
            clv = Decimal(str(row.total_spent)) / Decimal(str(lifespan_months))
            clv_values.append(float(clv))
            
            customer_clvs.append({
                "user_id": str(row.id),
                "clv": clv,
                "total_spent": Decimal(str(row.total_spent)),
                "transaction_count": row.transaction_count,
                "lifespan_months": lifespan_months
            })
        
        # Calculate statistics
        average_clv = Decimal(str(sum(clv_values) / len(clv_values)))
        sorted_clvs = sorted(clv_values)
        median_clv = Decimal(str(sorted_clvs[len(sorted_clvs) // 2]))
        
        # CLV by segment
        clv_by_segment = {}
        if not segment:  # Only calculate segments if not already filtered
            # High value customers (top 20%)
            high_value_threshold = sorted(clv_values, reverse=True)[int(len(clv_values) * 0.2)]
            high_value_clvs = [clv for clv in clv_values if clv >= high_value_threshold]
            clv_by_segment["high_value"] = Decimal(str(sum(high_value_clvs) / len(high_value_clvs))) if high_value_clvs else Decimal("0")
            
            # Medium value customers (middle 60%)
            medium_low = sorted(clv_values, reverse=True)[int(len(clv_values) * 0.8)]
            medium_clvs = [clv for clv in clv_values if medium_low <= clv < high_value_threshold]
            clv_by_segment["medium_value"] = Decimal(str(sum(medium_clvs) / len(medium_clvs))) if medium_clvs else Decimal("0")
            
            # Low value customers (bottom 20%)
            low_clvs = [clv for clv in clv_values if clv < medium_low]
            clv_by_segment["low_value"] = Decimal(str(sum(low_clvs) / len(low_clvs))) if low_clvs else Decimal("0")
        
        # CLV trend (monthly averages for last 12 months)
        clv_trend = []
        end_date = datetime.now()
        for i in range(12):
            month_start = end_date - timedelta(days=30 * (i + 1))
            month_end = end_date - timedelta(days=30 * i)
            
            month_customers = [
                c for c in customer_clvs 
                if month_start <= c.get("last_transaction", datetime.now()) <= month_end
            ]
            
            if month_customers:
                month_avg_clv = sum(float(c["clv"]) for c in month_customers) / len(month_customers)
                clv_trend.append({
                    "month": month_start.strftime("%Y-%m"),
                    "average_clv": round(month_avg_clv, 2),
                    "customer_count": len(month_customers)
                })
        
        clv_trend.reverse()  # Chronological order
        
        # Top customers by CLV
        top_customers = sorted(customer_clvs, key=lambda x: x["clv"], reverse=True)[:10]
        top_customers_formatted = [
            {
                "user_id": c["user_id"],
                "clv": float(c["clv"]),
                "total_spent": float(c["total_spent"]),
                "transaction_count": c["transaction_count"],
                "lifespan_months": round(c["lifespan_months"], 1)
            }
            for c in top_customers
        ]
        
        return CLVAnalysis(
            average_clv=average_clv,
            median_clv=median_clv,
            clv_by_segment=clv_by_segment,
            clv_trend=clv_trend,
            top_customers=top_customers_formatted
        )

    async def get_payment_success_analytics(
        self, 
        date_range: DateRange, 
        breakdown_by: List[str]
    ) -> PaymentAnalytics:
        """
        Get payment success analytics with breakdown by method and region.
        
        Requirements: 12.4, 12.5
        """
        # Base conditions
        date_conditions = and_(
            PaymentIntent.created_at >= date_range.start_date,
            PaymentIntent.created_at <= date_range.end_date
        )
        
        # Total payments
        total_query = select(func.count(PaymentIntent.id)).where(date_conditions)
        total_result = await self.db.execute(total_query)
        total_payments = total_result.scalar() or 0
        
        # Successful payments
        success_query = select(func.count(PaymentIntent.id)).where(
            and_(date_conditions, PaymentIntent.status == "succeeded")
        )
        success_result = await self.db.execute(success_query)
        successful_payments = success_result.scalar() or 0
        
        # Failed payments
        failed_query = select(func.count(PaymentIntent.id)).where(
            and_(date_conditions, PaymentIntent.status.in_(["failed", "canceled"]))
        )
        failed_result = await self.db.execute(failed_query)
        failed_payments = failed_result.scalar() or 0
        
        # Success rate
        success_rate = (successful_payments / total_payments * 100) if total_payments > 0 else 0.0
        
        # Total volume
        volume_query = select(
            func.sum(
                func.cast(
                    func.json_extract_path_text(PaymentIntent.amount_breakdown, 'total_amount'),
                    func.DECIMAL
                )
            )
        ).where(and_(date_conditions, PaymentIntent.status == "succeeded"))
        volume_result = await self.db.execute(volume_query)
        total_volume = Decimal(str(volume_result.scalar() or 0))
        
        # Average payment amount
        avg_amount = total_volume / successful_payments if successful_payments > 0 else Decimal("0")
        
        # Breakdown by payment method
        breakdown_by_method = {}
        if "payment_method" in breakdown_by:
            method_query = select(
                PaymentIntent.payment_method_type,
                func.count(PaymentIntent.id).label("total"),
                func.sum(
                    case(
                        (PaymentIntent.status == "succeeded", 1),
                        else_=0
                    )
                ).label("successful"),
                func.sum(
                    case(
                        (PaymentIntent.status == "succeeded", 
                         func.cast(
                             func.json_extract_path_text(PaymentIntent.amount_breakdown, 'total_amount'),
                             func.DECIMAL
                         )),
                        else_=0
                    )
                ).label("volume")
            ).where(date_conditions).group_by(PaymentIntent.payment_method_type)
            
            method_result = await self.db.execute(method_query)
            for row in method_result.all():
                method_type = row.payment_method_type or "unknown"
                breakdown_by_method[method_type] = {
                    "total_payments": row.total,
                    "successful_payments": row.successful,
                    "success_rate": (row.successful / row.total * 100) if row.total > 0 else 0.0,
                    "volume": float(row.volume or 0)
                }
        
        # Breakdown by country
        breakdown_by_country = {}
        if "country" in breakdown_by:
            country_query = select(
                User.country,
                func.count(PaymentIntent.id).label("total"),
                func.sum(
                    case(
                        (PaymentIntent.status == "succeeded", 1),
                        else_=0
                    )
                ).label("successful"),
                func.sum(
                    case(
                        (PaymentIntent.status == "succeeded", 
                         func.cast(
                             func.json_extract_path_text(PaymentIntent.amount_breakdown, 'total_amount'),
                             func.DECIMAL
                         )),
                        else_=0
                    )
                ).label("volume")
            ).join(User, PaymentIntent.user_id == User.id).where(
                date_conditions
            ).group_by(User.country)
            
            country_result = await self.db.execute(country_query)
            for row in country_result.all():
                country = row.country or "Unknown"
                breakdown_by_country[country] = {
                    "total_payments": row.total,
                    "successful_payments": row.successful,
                    "success_rate": (row.successful / row.total * 100) if row.total > 0 else 0.0,
                    "volume": float(row.volume or 0)
                }
        
        # Failure analysis
        failure_query = select(
            PaymentIntent.failure_reason,
            func.count(PaymentIntent.id).label("count")
        ).where(
            and_(date_conditions, PaymentIntent.status.in_(["failed", "canceled"]))
        ).group_by(PaymentIntent.failure_reason)
        
        failure_result = await self.db.execute(failure_query)
        failure_analysis = {
            row.failure_reason or "unknown": row.count
            for row in failure_result.all()
        }
        
        return PaymentAnalytics(
            total_payments=total_payments,
            successful_payments=successful_payments,
            failed_payments=failed_payments,
            success_rate=success_rate,
            total_volume=total_volume,
            average_payment_amount=avg_amount,
            breakdown_by_method=breakdown_by_method,
            breakdown_by_country=breakdown_by_country,
            failure_analysis=failure_analysis
        )

    async def generate_revenue_forecast(
        self, 
        forecast_months: int, 
        confidence_level: float = 0.95
    ) -> RevenueForecast:
        """
        Generate revenue forecast based on historical patterns.
        
        Requirements: 12.5
        """
        # Get historical revenue data for the last 24 months
        end_date = datetime.now()
        start_date = end_date - timedelta(days=730)  # ~24 months
        
        # Monthly revenue from subscriptions
        subscription_revenue_query = select(
            func.date_trunc('month', Subscription.created_at).label('month'),
            func.sum(Subscription.price).label('revenue'),
            func.count(Subscription.id).label('subscription_count')
        ).where(
            and_(
                Subscription.created_at >= start_date,
                Subscription.created_at <= end_date,
                Subscription.status.in_(["active", "paused", "canceled"])
            )
        ).group_by(
            func.date_trunc('month', Subscription.created_at)
        ).order_by(
            func.date_trunc('month', Subscription.created_at)
        )
        
        subscription_result = await self.db.execute(subscription_revenue_query)
        historical_data = []
        
        for row in subscription_result.all():
            historical_data.append({
                "month": row.month.strftime("%Y-%m"),
                "revenue": float(row.revenue or 0),
                "subscription_count": row.subscription_count,
                "date": row.month
            })
        
        if len(historical_data) < 3:
            # Not enough data for meaningful forecast
            return RevenueForecast(
                forecast_data=[],
                confidence_intervals={},
                growth_rate=0.0,
                seasonal_factors={}
            )
        
        # Calculate growth rate (simple linear trend)
        revenues = [d["revenue"] for d in historical_data]
        months = list(range(len(revenues)))
        
        # Simple linear regression for trend
        n = len(revenues)
        sum_x = sum(months)
        sum_y = sum(revenues)
        sum_xy = sum(x * y for x, y in zip(months, revenues))
        sum_x2 = sum(x * x for x in months)
        
        if n * sum_x2 - sum_x * sum_x != 0:
            slope = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x * sum_x)
            intercept = (sum_y - slope * sum_x) / n
        else:
            slope = 0
            intercept = sum_y / n if n > 0 else 0
        
        # Calculate growth rate as percentage
        avg_revenue = sum(revenues) / len(revenues) if revenues else 0
        growth_rate = (slope / avg_revenue * 100) if avg_revenue > 0 else 0.0
        
        # Generate forecast
        forecast_data = []
        last_month = historical_data[-1]["date"] if historical_data else end_date
        
        for i in range(1, forecast_months + 1):
            forecast_month = last_month + timedelta(days=30 * i)
            
            # Base forecast using trend
            base_forecast = intercept + slope * (len(revenues) + i)
            
            # Apply seasonal adjustment (simplified)
            month_num = forecast_month.month
            seasonal_factor = 1.0
            if month_num in [11, 12]:  # Holiday season boost
                seasonal_factor = 1.15
            elif month_num in [1, 2]:  # Post-holiday dip
                seasonal_factor = 0.9
            
            forecasted_revenue = max(base_forecast * seasonal_factor, 0)
            
            # Calculate confidence intervals (simplified)
            # In a real implementation, you'd use statistical methods
            margin_of_error = forecasted_revenue * 0.2  # 20% margin
            lower_bound = max(forecasted_revenue - margin_of_error, 0)
            upper_bound = forecasted_revenue + margin_of_error
            
            forecast_data.append({
                "month": forecast_month.strftime("%Y-%m"),
                "forecasted_revenue": round(forecasted_revenue, 2),
                "lower_bound": round(lower_bound, 2),
                "upper_bound": round(upper_bound, 2),
                "confidence_level": confidence_level
            })
        
        # Confidence intervals summary
        confidence_intervals = {
            "method": "linear_trend_with_seasonal_adjustment",
            "confidence_level": confidence_level,
            "margin_of_error_percent": 20.0
        }
        
        # Seasonal factors
        seasonal_factors = {
            "january": 0.9,
            "february": 0.9,
            "march": 1.0,
            "april": 1.0,
            "may": 1.0,
            "june": 1.0,
            "july": 1.0,
            "august": 1.0,
            "september": 1.0,
            "october": 1.0,
            "november": 1.15,
            "december": 1.15
        }
        
        return RevenueForecast(
            forecast_data=forecast_data,
            confidence_intervals=confidence_intervals,
            growth_rate=growth_rate,
            seasonal_factors=seasonal_factors
        )

    async def get_live_metrics_display(
        self, 
        refresh_interval_seconds: int = 30
    ) -> Dict[str, Any]:
        """
        Create live metrics display for subscription revenue and performance.
        
        Requirements: 12.6
        """
        current_time = datetime.now()
        
        # Real-time active subscriptions
        active_subs_query = select(func.count(Subscription.id)).where(
            Subscription.status == "active"
        )
        active_subs_result = await self.db.execute(active_subs_query)
        active_subscriptions = active_subs_result.scalar() or 0
        
        # Real-time MRR
        mrr_query = select(func.sum(
            case(
                (Subscription.billing_cycle == "monthly", Subscription.price),
                (Subscription.billing_cycle == "yearly", Subscription.price / 12),
                else_=Subscription.price
            )
        )).where(Subscription.status == "active")
        mrr_result = await self.db.execute(mrr_query)
        current_mrr = float(mrr_result.scalar() or 0)
        
        # Today's new subscriptions
        today_start = current_time.replace(hour=0, minute=0, second=0, microsecond=0)
        today_subs_query = select(func.count(Subscription.id)).where(
            and_(
                Subscription.created_at >= today_start,
                Subscription.created_at <= current_time
            )
        )
        today_subs_result = await self.db.execute(today_subs_query)
        todays_new_subscriptions = today_subs_result.scalar() or 0
        
        # Today's revenue
        today_revenue_query = select(func.sum(
            func.cast(
                func.json_extract_path_text(PaymentIntent.amount_breakdown, 'total_amount'),
                func.DECIMAL
            )
        )).where(
            and_(
                PaymentIntent.created_at >= today_start,
                PaymentIntent.created_at <= current_time,
                PaymentIntent.status == "succeeded"
            )
        )
        today_revenue_result = await self.db.execute(today_revenue_query)
        todays_revenue = float(today_revenue_result.scalar() or 0)
        
        # Recent payment success rate (last hour)
        hour_ago = current_time - timedelta(hours=1)
        recent_payments_query = select(func.count(PaymentIntent.id)).where(
            PaymentIntent.created_at >= hour_ago
        )
        recent_payments_result = await self.db.execute(recent_payments_query)
        recent_total_payments = recent_payments_result.scalar() or 0
        
        recent_success_query = select(func.count(PaymentIntent.id)).where(
            and_(
                PaymentIntent.created_at >= hour_ago,
                PaymentIntent.status == "succeeded"
            )
        )
        recent_success_result = await self.db.execute(recent_success_query)
        recent_successful_payments = recent_success_result.scalar() or 0
        
        recent_success_rate = (
            recent_successful_payments / recent_total_payments * 100
        ) if recent_total_payments > 0 else 0.0
        
        # Growth metrics (compared to yesterday)
        yesterday_start = today_start - timedelta(days=1)
        yesterday_end = today_start
        
        yesterday_subs_query = select(func.count(Subscription.id)).where(
            and_(
                Subscription.created_at >= yesterday_start,
                Subscription.created_at < yesterday_end
            )
        )
        yesterday_subs_result = await self.db.execute(yesterday_subs_query)
        yesterday_new_subs = yesterday_subs_result.scalar() or 0
        
        subscription_growth = (
            ((todays_new_subscriptions - yesterday_new_subs) / yesterday_new_subs * 100)
            if yesterday_new_subs > 0 else 0.0
        )
        
        return {
            "timestamp": current_time.isoformat(),
            "refresh_interval_seconds": refresh_interval_seconds,
            "live_metrics": {
                "active_subscriptions": active_subscriptions,
                "monthly_recurring_revenue": round(current_mrr, 2),
                "todays_new_subscriptions": todays_new_subscriptions,
                "todays_revenue": round(todays_revenue, 2),
                "recent_payment_success_rate": round(recent_success_rate, 2),
                "subscription_growth_vs_yesterday": round(subscription_growth, 2)
            },
            "performance_indicators": {
                "mrr_trend": "up" if current_mrr > 0 else "stable",
                "subscription_trend": "up" if subscription_growth > 0 else "down" if subscription_growth < 0 else "stable",
                "payment_health": "good" if recent_success_rate > 95 else "warning" if recent_success_rate > 85 else "critical"
            }
        }

    async def get_filtered_analytics_with_date_range(
        self, 
        start_date: datetime, 
        end_date: datetime, 
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Add custom date range filtering using actual historical data.
        
        Requirements: 12.7
        """
        if filters is None:
            filters = {}
        
        date_range = DateRange(start_date=start_date, end_date=end_date)
        analytics_filters = AnalyticsFilters(
            user_segment=filters.get("user_segment"),
            subscription_status=filters.get("subscription_status"),
            payment_method=filters.get("payment_method"),
            currency=filters.get("currency"),
            country=filters.get("country"),
            category=filters.get("category")
        )
        
        # Get comprehensive metrics
        subscription_metrics = await self.get_subscription_metrics(date_range, analytics_filters)
        payment_analytics = await self.get_payment_success_analytics(
            date_range, 
            ["payment_method", "country"]
        )
        
        # Get daily breakdown for the date range
        daily_breakdown = await self._get_daily_breakdown(date_range, analytics_filters)
        
        # Get top performing segments
        top_segments = await self._get_top_performing_segments(date_range, analytics_filters)
        
        return {
            "date_range": {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "days": (end_date - start_date).days
            },
            "applied_filters": filters,
            "subscription_metrics": {
                "total_active_subscriptions": subscription_metrics.total_active_subscriptions,
                "new_subscriptions": subscription_metrics.new_subscriptions,
                "canceled_subscriptions": subscription_metrics.canceled_subscriptions,
                "total_revenue": float(subscription_metrics.total_revenue),
                "average_subscription_value": float(subscription_metrics.average_subscription_value),
                "churn_rate": subscription_metrics.churn_rate,
                "conversion_rate": subscription_metrics.conversion_rate,
                "plan_breakdown": subscription_metrics.plan_breakdown,
                "geographic_breakdown": subscription_metrics.geographic_breakdown
            },
            "payment_analytics": {
                "total_payments": payment_analytics.total_payments,
                "successful_payments": payment_analytics.successful_payments,
                "success_rate": payment_analytics.success_rate,
                "total_volume": float(payment_analytics.total_volume),
                "breakdown_by_method": payment_analytics.breakdown_by_method,
                "breakdown_by_country": payment_analytics.breakdown_by_country,
                "failure_analysis": payment_analytics.failure_analysis
            },
            "daily_breakdown": daily_breakdown,
            "top_segments": top_segments
        }

    async def monitor_payment_success_rate_and_failures(
        self, 
        monitoring_window_hours: int = 24
    ) -> Dict[str, Any]:
        """
        Implement payment success rate monitoring and failure analysis.
        
        Requirements: 12.7
        """
        current_time = datetime.now()
        window_start = current_time - timedelta(hours=monitoring_window_hours)
        
        # Overall success rate in monitoring window
        total_query = select(func.count(PaymentIntent.id)).where(
            PaymentIntent.created_at >= window_start
        )
        total_result = await self.db.execute(total_query)
        total_payments = total_result.scalar() or 0
        
        success_query = select(func.count(PaymentIntent.id)).where(
            and_(
                PaymentIntent.created_at >= window_start,
                PaymentIntent.status == "succeeded"
            )
        )
        success_result = await self.db.execute(success_query)
        successful_payments = success_result.scalar() or 0
        
        overall_success_rate = (
            successful_payments / total_payments * 100
        ) if total_payments > 0 else 0.0
        
        # Hourly breakdown
        hourly_breakdown = []
        for i in range(monitoring_window_hours):
            hour_start = window_start + timedelta(hours=i)
            hour_end = hour_start + timedelta(hours=1)
            
            hour_total_query = select(func.count(PaymentIntent.id)).where(
                and_(
                    PaymentIntent.created_at >= hour_start,
                    PaymentIntent.created_at < hour_end
                )
            )
            hour_total_result = await self.db.execute(hour_total_query)
            hour_total = hour_total_result.scalar() or 0
            
            hour_success_query = select(func.count(PaymentIntent.id)).where(
                and_(
                    PaymentIntent.created_at >= hour_start,
                    PaymentIntent.created_at < hour_end,
                    PaymentIntent.status == "succeeded"
                )
            )
            hour_success_result = await self.db.execute(hour_success_query)
            hour_successful = hour_success_result.scalar() or 0
            
            hour_success_rate = (
                hour_successful / hour_total * 100
            ) if hour_total > 0 else 0.0
            
            hourly_breakdown.append({
                "hour": hour_start.strftime("%Y-%m-%d %H:00"),
                "total_payments": hour_total,
                "successful_payments": hour_successful,
                "success_rate": round(hour_success_rate, 2)
            })
        
        # Failure analysis
        failure_query = select(
            PaymentIntent.failure_reason,
            PaymentIntent.payment_method_type,
            func.count(PaymentIntent.id).label("count")
        ).where(
            and_(
                PaymentIntent.created_at >= window_start,
                PaymentIntent.status.in_(["failed", "canceled"])
            )
        ).group_by(PaymentIntent.failure_reason, PaymentIntent.payment_method_type)
        
        failure_result = await self.db.execute(failure_query)
        failure_breakdown = []
        
        for row in failure_result.all():
            failure_breakdown.append({
                "failure_reason": row.failure_reason or "unknown",
                "payment_method": row.payment_method_type or "unknown",
                "count": row.count
            })
        
        # Alert conditions
        alerts = []
        if overall_success_rate < 85:
            alerts.append({
                "type": "critical",
                "message": f"Payment success rate critically low: {overall_success_rate:.1f}%",
                "threshold": 85,
                "current_value": overall_success_rate
            })
        elif overall_success_rate < 95:
            alerts.append({
                "type": "warning",
                "message": f"Payment success rate below optimal: {overall_success_rate:.1f}%",
                "threshold": 95,
                "current_value": overall_success_rate
            })
        
        # Check for sudden drops
        recent_hours = hourly_breakdown[-3:] if len(hourly_breakdown) >= 3 else hourly_breakdown
        if recent_hours:
            recent_avg = sum(h["success_rate"] for h in recent_hours) / len(recent_hours)
            if recent_avg < overall_success_rate - 10:
                alerts.append({
                    "type": "warning",
                    "message": f"Recent payment success rate drop detected: {recent_avg:.1f}%",
                    "threshold": overall_success_rate - 10,
                    "current_value": recent_avg
                })
        
        return {
            "monitoring_window_hours": monitoring_window_hours,
            "timestamp": current_time.isoformat(),
            "overall_metrics": {
                "total_payments": total_payments,
                "successful_payments": successful_payments,
                "failed_payments": total_payments - successful_payments,
                "success_rate": round(overall_success_rate, 2)
            },
            "hourly_breakdown": hourly_breakdown,
            "failure_analysis": failure_breakdown,
            "alerts": alerts,
            "recommendations": self._generate_payment_recommendations(failure_breakdown, overall_success_rate)
        }

    async def _get_daily_breakdown(
        self, 
        date_range: DateRange, 
        filters: AnalyticsFilters
    ) -> List[Dict[str, Any]]:
        """Get daily breakdown of metrics within date range."""
        daily_data = []
        current_date = date_range.start_date.date()
        end_date = date_range.end_date.date()
        
        while current_date <= end_date:
            day_start = datetime.combine(current_date, datetime.min.time())
            day_end = day_start + timedelta(days=1)
            
            # Daily subscriptions
            daily_subs_query = select(func.count(Subscription.id)).where(
                and_(
                    Subscription.created_at >= day_start,
                    Subscription.created_at < day_end
                )
            )
            daily_subs_result = await self.db.execute(daily_subs_query)
            daily_subscriptions = daily_subs_result.scalar() or 0
            
            # Daily revenue
            daily_revenue_query = select(func.sum(
                func.cast(
                    func.json_extract_path_text(PaymentIntent.amount_breakdown, 'total_amount'),
                    func.DECIMAL
                )
            )).where(
                and_(
                    PaymentIntent.created_at >= day_start,
                    PaymentIntent.created_at < day_end,
                    PaymentIntent.status == "succeeded"
                )
            )
            daily_revenue_result = await self.db.execute(daily_revenue_query)
            daily_revenue = float(daily_revenue_result.scalar() or 0)
            
            daily_data.append({
                "date": current_date.isoformat(),
                "new_subscriptions": daily_subscriptions,
                "revenue": daily_revenue
            })
            
            current_date += timedelta(days=1)
        
        return daily_data

    async def _get_top_performing_segments(
        self, 
        date_range: DateRange, 
        filters: AnalyticsFilters
    ) -> Dict[str, Any]:
        """Get top performing customer segments."""
        # Top countries by revenue
        country_query = select(
            User.country,
            func.count(Subscription.id).label("subscription_count"),
            func.sum(Subscription.price).label("total_revenue")
        ).join(User, Subscription.user_id == User.id).where(
            and_(
                Subscription.created_at >= date_range.start_date,
                Subscription.created_at <= date_range.end_date
            )
        ).group_by(User.country).order_by(desc("total_revenue")).limit(5)
        
        country_result = await self.db.execute(country_query)
        top_countries = [
            {
                "country": row.country or "Unknown",
                "subscription_count": row.subscription_count,
                "total_revenue": float(row.total_revenue or 0)
            }
            for row in country_result.all()
        ]
        
        # Top subscription plans
        plan_query = select(
            Subscription.plan_id,
            func.count(Subscription.id).label("subscription_count"),
            func.sum(Subscription.price).label("total_revenue")
        ).where(
            and_(
                Subscription.created_at >= date_range.start_date,
                Subscription.created_at <= date_range.end_date
            )
        ).group_by(Subscription.plan_id).order_by(desc("total_revenue")).limit(5)
        
        plan_result = await self.db.execute(plan_query)
        top_plans = [
            {
                "plan_id": row.plan_id,
                "subscription_count": row.subscription_count,
                "total_revenue": float(row.total_revenue or 0)
            }
            for row in plan_result.all()
        ]
        
        return {
            "top_countries": top_countries,
            "top_plans": top_plans
        }

    def _generate_payment_recommendations(
        self, 
        failure_breakdown: List[Dict[str, Any]], 
        success_rate: float
    ) -> List[str]:
        """Generate recommendations based on payment failure analysis."""
        recommendations = []
        
        if success_rate < 85:
            recommendations.append("Critical: Investigate payment processor issues immediately")
        
        # Analyze failure patterns
        failure_counts = {}
        for failure in failure_breakdown:
            reason = failure["failure_reason"]
            failure_counts[reason] = failure_counts.get(reason, 0) + failure["count"]
        
        if failure_counts:
            top_failure = max(failure_counts.items(), key=lambda x: x[1])
            if top_failure[0] == "insufficient_funds":
                recommendations.append("Consider implementing payment retry logic for insufficient funds")
            elif top_failure[0] == "card_declined":
                recommendations.append("Review fraud detection settings - may be too strict")
            elif "authentication" in top_failure[0].lower():
                recommendations.append("Optimize 3D Secure flow to reduce authentication failures")
        
        if success_rate < 95:
            recommendations.append("Monitor payment method performance and consider alternative providers")
        
        return recommendations

    # Legacy methods - keeping for backward compatibility
    async def get_dashboard_data(self, user_id: str, user_role: str, start_date: datetime, end_date: datetime = None, filters: dict = None) -> dict:
        """Get dashboard analytics data based on user role with filters."""
        if end_date is None:
            end_date = datetime.now()
        if filters is None:
            filters = {}
            
        if user_role in ["Admin", "SuperAdmin"]:
            return await self._get_admin_dashboard_data(start_date, end_date, filters)
        elif user_role == "Supplier":
            return await self._get_supplier_dashboard_data(user_id, start_date, end_date, filters)
        else:
            return await self._get_customer_dashboard_data(user_id, start_date, end_date, filters)

    async def _get_admin_dashboard_data(self, start_date: datetime, end_date: datetime, filters: dict) -> dict:
        """Get admin dashboard data with filters."""
        # Build base conditions
        date_conditions = and_(
            Order.created_at >= start_date,
            Order.created_at <= end_date
        )
        
        # Apply filters to queries
        filter_conditions = []
        if filters.get('order_status'):
            filter_conditions.append(Order.status == filters['order_status'])
        
        # For category and product filters, we need to join with OrderItem and Product
        needs_product_join = filters.get('category') or filters.get('product')
        
        # Total sales
        if needs_product_join:
            sales_query = select(func.sum(OrderItem.total_price)).join(
                Order, OrderItem.order_id == Order.id
            ).join(
                ProductVariant, OrderItem.variant_id == ProductVariant.id
            ).join(
                Product, ProductVariant.product_id == Product.id
            ).where(date_conditions)
            
            if filters.get('category'):
                # Join with Category table to filter by category name
                sales_query = sales_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
            if filters.get('product'):
                sales_query = sales_query.where(Product.name.ilike(f"%{filters['product']}%"))
            if filter_conditions:
                sales_query = sales_query.where(and_(*filter_conditions))
        else:
            sales_query = select(func.sum(Order.total_amount)).where(date_conditions)
            if filter_conditions:
                sales_query = sales_query.where(and_(*filter_conditions))
                
        sales_result = await self.db.execute(sales_query)
        total_sales = sales_result.scalar() or 0

        # Total orders
        orders_query = select(func.count(Order.id.distinct())).where(date_conditions)
        if filter_conditions:
            orders_query = orders_query.where(and_(*filter_conditions))
        if needs_product_join:
            orders_query = select(func.count(Order.id.distinct())).join(
                OrderItem, Order.id == OrderItem.order_id
            ).join(
                ProductVariant, OrderItem.variant_id == ProductVariant.id
            ).join(
                Product, ProductVariant.product_id == Product.id
            ).where(date_conditions)
            if filters.get('category'):
                orders_query = orders_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
            if filters.get('product'):
                orders_query = orders_query.where(Product.name.ilike(f"%{filters['product']}%"))
            if filter_conditions:
                orders_query = orders_query.where(and_(*filter_conditions))
                
        orders_result = await self.db.execute(orders_query)
        total_orders = orders_result.scalar() or 0

        # Total users (with user segment filter)
        user_date_conditions = and_(
            User.created_at >= start_date,
            User.created_at <= end_date
        )
        users_query = select(func.count(User.id)).where(user_date_conditions)
        
        # Apply user segment filter
        if filters.get('user_segment'):
            if filters['user_segment'] == 'new':
                # Users with no orders
                users_query = select(func.count(User.id)).where(
                    and_(
                        user_date_conditions,
                        ~User.id.in_(select(Order.user_id.distinct()))
                    )
                )
            elif filters['user_segment'] == 'returning':
                # Users with more than one order
                users_query = select(func.count(User.id.distinct())).join(
                    Order, User.id == Order.user_id
                ).where(user_date_conditions).group_by(User.id).having(
                    func.count(Order.id) > 1
                )
            elif filters['user_segment'] == 'vip':
                # Users with total spending > threshold (e.g., $1000)
                users_query = select(func.count(User.id.distinct())).join(
                    Order, User.id == Order.user_id
                ).where(user_date_conditions).group_by(User.id).having(
                    func.sum(Order.total_amount) > 1000
                )
                
        users_result = await self.db.execute(users_query)
        total_users = users_result.scalar() or 0

        # Total products
        products_query = select(func.count(Product.id))
        if filters.get('category'):
            products_query = products_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        products_result = await self.db.execute(products_query)
        total_products = products_result.scalar() or 0

        # Get sales trend and top products with filters
        sales_trend = await self.get_sales_trend(
            user_id="", 
            user_role="Admin", 
            days=(end_date - start_date).days,
            filters=filters,
            start_date=start_date,
            end_date=end_date
        )
        top_products = await self.get_top_products(
            user_id="", 
            user_role="Admin", 
            limit=5,
            filters=filters
        )

        # Order status distribution
        order_status_query = select(
            Order.status,
            func.count(Order.id).label("count")
        ).where(date_conditions)
        
        if needs_product_join:
            order_status_query = select(
                Order.status,
                func.count(Order.id.distinct()).label("count")
            ).join(
                OrderItem, Order.id == OrderItem.order_id
            ).join(
                ProductVariant, OrderItem.variant_id == ProductVariant.id
            ).join(
                Product, ProductVariant.product_id == Product.id
            ).where(date_conditions)
            if filters.get('category'):
                order_status_query = order_status_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
            if filters.get('product'):
                order_status_query = order_status_query.where(Product.name.ilike(f"%{filters['product']}%"))
        
        if filter_conditions:
            order_status_query = order_status_query.where(and_(*filter_conditions))
            
        order_status_query = order_status_query.group_by(Order.status)
        order_status_result = await self.db.execute(order_status_query)
        order_status_distribution = {
            row.status: row.count for row in order_status_result.all()}

        # User growth
        user_growth_query = select(
            func.date(User.created_at).label("date"),
            func.count(User.id).label("new_users")
        ).where(user_date_conditions).group_by(
            func.date(User.created_at)
        ).order_by(
            func.date(User.created_at)
        )
        user_growth_result = await self.db.execute(user_growth_query)
        user_growth = [{
            "date": row.date.strftime("%Y-%m-%d") if hasattr(row.date, 'strftime') else str(row.date),
            "new_users": row.new_users
        } for row in user_growth_result.all()]

        # Conversion rate (simplified: unique purchasing users / total users)
        purchasing_users_query = select(func.count(Order.user_id.distinct())).where(date_conditions)
        if filter_conditions:
            purchasing_users_query = purchasing_users_query.where(and_(*filter_conditions))
        purchasing_users_result = await self.db.execute(purchasing_users_query)
        unique_purchasing_users = purchasing_users_result.scalar() or 0

        conversion_rate = (unique_purchasing_users /
                           total_users * 100) if total_users > 0 else 0.0

        return {
            "total_sales": float(total_sales),
            "total_orders": total_orders,
            "total_users": total_users,
            "total_products": total_products,
            "conversion_rate": round(conversion_rate, 2),
            "average_order_value": float(total_sales / total_orders) if total_orders > 0 else 0,
            "top_products": top_products,
            "sales_trend": sales_trend,
            "order_status_distribution": order_status_distribution,
            "user_growth": user_growth
        }

    async def _get_supplier_dashboard_data(self, user_id: str, start_date: datetime, end_date: datetime, filters: dict) -> dict:
        """Get supplier dashboard data with filters."""
        # Get supplier's products
        products_query = select(Product).where(Product.supplier_id == user_id)
        if filters.get('category'):
            products_query = products_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        products_result = await self.db.execute(products_query)
        products = products_result.scalars().all()

        # Calculate total sales and orders for the supplier
        order_item_alias = aliased(OrderItem)
        product_variant_alias = aliased(ProductVariant)

        date_conditions = and_(
            Order.created_at >= start_date,
            Order.created_at <= end_date
        )

        supplier_orders_query = select(Order).join(
            order_item_alias, Order.id == order_item_alias.order_id
        ).join(
            product_variant_alias, order_item_alias.variant_id == product_variant_alias.id
        ).join(
            Product, product_variant_alias.product_id == Product.id
        ).where(
            and_(
                Product.supplier_id == user_id,
                date_conditions
            )
        )
        
        if filters.get('category'):
            supplier_orders_query = supplier_orders_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        if filters.get('order_status'):
            supplier_orders_query = supplier_orders_query.where(Order.status == filters['order_status'])
            
        supplier_orders_query = supplier_orders_query.distinct()

        supplier_orders_result = await self.db.execute(supplier_orders_query)
        supplier_orders = supplier_orders_result.scalars().all()

        total_sales = sum(order.total_amount for order in supplier_orders)
        total_orders = len(supplier_orders)

        # Calculate average rating for supplier's products
        reviews_query = select(func.avg(Review.rating)).join(
            Product, Review.product_id == Product.id
        ).where(
            Product.supplier_id == user_id
        )
        if filters.get('category'):
            reviews_query = reviews_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        average_rating_result = await self.db.execute(reviews_query)
        average_rating = average_rating_result.scalar() or 0.0

        return {
            "total_products": len(products),
            "total_sales": float(total_sales),
            "total_orders": total_orders,
            "average_rating": round(average_rating, 2),
            "products": [
                {
                    "id": str(p.id),
                    "name": p.name,
                    "rating": p.rating,
                    "review_count": p.review_count
                }
                for p in products[:5]
            ]
        }

    async def _get_customer_dashboard_data(self, user_id: str, start_date: datetime, end_date: datetime, filters: dict) -> dict:
        """Get customer dashboard data with filters."""
        # Get customer's orders
        date_conditions = and_(
            Order.user_id == user_id,
            Order.created_at >= start_date,
            Order.created_at <= end_date
        )
        orders_query = select(Order).where(date_conditions)
        
        if filters.get('order_status'):
            orders_query = orders_query.where(Order.status == filters['order_status'])
            
        orders_result = await self.db.execute(orders_query)
        orders = orders_result.scalars().all()

        total_spent = sum(order.total_amount for order in orders)

        return {
            "total_orders": len(orders),
            "total_spent": float(total_spent),
            "average_order_value": float(total_spent / len(orders)) if orders else 0,
            "recent_orders": [
                {
                    "id": str(order.id),
                    "status": order.status,
                    "total_amount": order.total_amount,
                    "created_at": order.created_at.isoformat()
                }
                for order in orders[:5]
            ]
        }

    async def get_sales_trend(self, user_id: str, user_role: str, days: int = 30, filters: dict = None, start_date: datetime = None, end_date: datetime = None) -> List[dict]:
        """Get sales trend data grouped by date with filters."""
        if filters is None:
            filters = {}
            
        # Use provided dates if available, otherwise calculate from days
        if end_date is None:
            end_date = datetime.now()
        if start_date is None:
            start_date = end_date - timedelta(days=days)

        date_conditions = and_(
            Order.created_at >= start_date,
            Order.created_at <= end_date
        )

        # Check if we need product joins for filtering
        needs_product_join = filters.get('category') or filters.get('product') or user_role == "Supplier"

        if needs_product_join:
            # Use OrderItem-based query for product filtering
            query = select(
                func.date(Order.created_at).label("date"),
                func.sum(OrderItem.total_price).label("sales"),
                func.count(Order.id.distinct()).label("orders")
            ).join(
                OrderItem, Order.id == OrderItem.order_id
            ).join(
                ProductVariant, OrderItem.variant_id == ProductVariant.id
            ).join(
                Product, ProductVariant.product_id == Product.id
            ).where(date_conditions)
            
            if user_role == "Supplier":
                query = query.where(Product.supplier_id == user_id)
            if filters.get('category'):
                query = query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
            if filters.get('product'):
                query = query.where(Product.name.ilike(f"%{filters['product']}%"))
            if filters.get('order_status'):
                query = query.where(Order.status == filters['order_status'])
                
            query = query.group_by(
                func.date(Order.created_at)
            ).order_by(
                func.date(Order.created_at)
            )
        else:
            # Use simple Order-based query
            query = select(
                func.date(Order.created_at).label("date"),
                func.sum(Order.total_amount).label("sales"),
                func.count(Order.id).label("orders")
            ).where(date_conditions)
            
            if filters.get('order_status'):
                query = query.where(Order.status == filters['order_status'])
                
            query = query.group_by(
                func.date(Order.created_at)
            ).order_by(
                func.date(Order.created_at)
            )

        result = await self.db.execute(query)
        trend_data = []
        for row in result.all():
            trend_data.append({
                "date": row.date.strftime("%Y-%m-%d") if hasattr(row.date, 'strftime') else str(row.date),
                "sales": float(row.sales or 0),
                "orders": int(row.orders or 0)
            })

        return trend_data

    async def get_top_products(self, user_id: str, user_role: str, limit: int = 10, filters: dict = None) -> List[dict]:
        """Get top performing products by revenue with filters."""
        if filters is None:
            filters = {}
            
        # Get the most recent orders to ensure fresh data
        from sqlalchemy.orm import selectinload
        
        query = select(
            Product.id,
            Product.name,
            func.sum(OrderItem.quantity).label("total_sales_quantity"),
            func.sum(OrderItem.total_price).label("total_revenue")
        ).join(
            ProductVariant, Product.id == ProductVariant.product_id
        ).join(
            OrderItem, ProductVariant.id == OrderItem.variant_id
        ).join(
            Order, OrderItem.order_id == Order.id
        ).where(
            Order.status.in_(['confirmed', 'shipped', 'delivered', 'processing'])
        )
        
        if user_role == "Supplier":
            query = query.where(Product.supplier_id == user_id)
        if filters.get('category'):
            query = query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        if filters.get('product'):
            query = query.where(Product.name.ilike(f"%{filters['product']}%"))
        if filters.get('order_status'):
            query = query.where(Order.status == filters['order_status'])
            
        query = query.group_by(
            Product.id,
            Product.name
        ).order_by(
            func.sum(OrderItem.total_price).desc()
        ).limit(limit)

        result = await self.db.execute(query)
        products_data = []
        for row in result.all():
            # Get product image from first variant
            image_query = select(ProductVariant).where(
                ProductVariant.product_id == row.id
            ).options(selectinload(ProductVariant.images)).limit(1)
            variant_result = await self.db.execute(image_query)
            variant = variant_result.scalar_one_or_none()
            
            image_url = None
            if variant and variant.images:
                primary_image = next((img for img in variant.images if img.is_primary), None)
                image_url = primary_image.url if primary_image else (variant.images[0].url if variant.images else None)
            
            products_data.append({
                "id": str(row.id),
                "name": row.name,
                "sales": int(row.total_sales_quantity or 0),
                "revenue": float(row.total_revenue or 0),
                "image_url": image_url
            })

        return products_data

    async def export_data(self, data: dict, format: str, export_type: str) -> Tuple[bytes, str, str]:
        """
        Export analytics data to CSV or Excel format.
        
        Args:
            data: The analytics data dictionary
            format: 'csv' or 'xlsx'
            export_type: Type of export (e.g., 'dashboard')
            
        Returns:
            Tuple of (file_content, content_type, filename)
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format == 'csv':
            return await self._export_csv(data, export_type, timestamp)
        elif format == 'xlsx':
            return await self._export_excel(data, export_type, timestamp)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    async def _export_csv(self, data: dict, export_type: str, timestamp: str) -> Tuple[bytes, str, str]:
        """Export data as CSV."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write summary section
        writer.writerow(['Analytics Summary'])
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Sales', f"${data.get('total_sales', 0):.2f}"])
        writer.writerow(['Total Orders', data.get('total_orders', 0)])
        writer.writerow(['Total Users', data.get('total_users', 0)])
        writer.writerow(['Total Products', data.get('total_products', 0)])
        writer.writerow(['Conversion Rate', f"{data.get('conversion_rate', 0):.2f}%"])
        writer.writerow(['Average Order Value', f"${data.get('average_order_value', 0):.2f}"])
        writer.writerow([])
        
        # Write sales trend section
        writer.writerow(['Sales Trend'])
        writer.writerow(['Date', 'Sales', 'Orders'])
        for item in data.get('sales_trend', []):
            writer.writerow([
                item.get('date', ''),
                f"${item.get('sales', 0):.2f}",
                item.get('orders', 0)
            ])
        writer.writerow([])
        
        # Write top products section
        writer.writerow(['Top Products'])
        writer.writerow(['Product Name', 'Units Sold', 'Revenue'])
        for product in data.get('top_products', []):
            writer.writerow([
                product.get('name', ''),
                product.get('sales', 0),
                f"${product.get('revenue', 0):.2f}"
            ])
        writer.writerow([])
        
        # Write order status distribution
        writer.writerow(['Order Status Distribution'])
        writer.writerow(['Status', 'Count'])
        for status, count in data.get('order_status_distribution', {}).items():
            writer.writerow([status, count])
        
        # Convert to bytes
        content = output.getvalue().encode('utf-8')
        filename = f"analytics_{export_type}_{timestamp}.csv"
        
        return content, 'text/csv', filename
    
    async def _export_excel(self, data: dict, export_type: str, timestamp: str) -> Tuple[bytes, str, str]:
        """Export data as Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Summary"
        
        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=14)
        
        # Write summary section
        row = 1
        ws.cell(row=row, column=1, value='Analytics Summary').font = title_font
        row += 1
        
        ws.cell(row=row, column=1, value='Metric').font = header_font
        ws.cell(row=row, column=2, value='Value').font = header_font
        row += 1
        
        summary_data = [
            ('Total Sales', f"${data.get('total_sales', 0):.2f}"),
            ('Total Orders', data.get('total_orders', 0)),
            ('Total Users', data.get('total_users', 0)),
            ('Total Products', data.get('total_products', 0)),
            ('Conversion Rate', f"{data.get('conversion_rate', 0):.2f}%"),
            ('Average Order Value', f"${data.get('average_order_value', 0):.2f}")
        ]
        
        for metric, value in summary_data:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 2
        
        # Write sales trend section
        ws.cell(row=row, column=1, value='Sales Trend').font = title_font
        row += 1
        
        ws.cell(row=row, column=1, value='Date').font = header_font
        ws.cell(row=row, column=2, value='Sales').font = header_font
        ws.cell(row=row, column=3, value='Orders').font = header_font
        row += 1
        
        for item in data.get('sales_trend', []):
            ws.cell(row=row, column=1, value=item.get('date', ''))
            ws.cell(row=row, column=2, value=f"${item.get('sales', 0):.2f}")
            ws.cell(row=row, column=3, value=item.get('orders', 0))
            row += 1
        
        row += 2
        
        # Write top products section
        ws.cell(row=row, column=1, value='Top Products').font = title_font
        row += 1
        
        ws.cell(row=row, column=1, value='Product Name').font = header_font
        ws.cell(row=row, column=2, value='Units Sold').font = header_font
        ws.cell(row=row, column=3, value='Revenue').font = header_font
        row += 1
        
        for product in data.get('top_products', []):
            ws.cell(row=row, column=1, value=product.get('name', ''))
            ws.cell(row=row, column=2, value=product.get('sales', 0))
            ws.cell(row=row, column=3, value=f"${product.get('revenue', 0):.2f}")
            row += 1
        
        row += 2
        
        # Write order status distribution
        ws.cell(row=row, column=1, value='Order Status Distribution').font = title_font
        row += 1
        
        ws.cell(row=row, column=1, value='Status').font = header_font
        ws.cell(row=row, column=2, value='Count').font = header_font
        row += 1
        
        for status, count in data.get('order_status_distribution', {}).items():
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        filename = f"analytics_{export_type}_{timestamp}.xlsx"
        
        return content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename

    async def get_recent_activity(self, limit: int = 100, since: datetime = None) -> List[dict]:
        """Get recent activity logs with user details and timestamps."""
        from models.activity_log import ActivityLog
        from sqlalchemy.orm import selectinload
        
        query = select(ActivityLog).options(
            selectinload(ActivityLog.user)
        ).order_by(ActivityLog.created_at.desc())
        
        if since:
            query = query.where(ActivityLog.created_at >= since)
        
        query = query.limit(limit)
        
        result = await self.db.execute(query)
        activities = result.scalars().all()
        
        return [
            {
                "id": str(activity.id),
                "user_id": str(activity.user_id) if activity.user_id else None,
                "user_name": activity.user.full_name if activity.user else "System",
                "action_type": activity.action_type,
                "description": activity.description,
                "metadata": activity.meta_data,  # Use meta_data column name
                "created_at": activity.created_at.isoformat() if activity.created_at else None
            }
            for activity in activities
        ]

    async def get_dashboard_data(self, user_id: str, user_role: str, start_date: datetime, end_date: datetime = None, filters: dict = None) -> dict:
        """Get dashboard analytics data based on user role with filters."""
        if end_date is None:
            end_date = datetime.now()
        if filters is None:
            filters = {}
            
        if user_role in ["Admin", "SuperAdmin"]:
            return await self._get_admin_dashboard_data(start_date, end_date, filters)
        elif user_role == "Supplier":
            return await self._get_supplier_dashboard_data(user_id, start_date, end_date, filters)
        else:
            return await self._get_customer_dashboard_data(user_id, start_date, end_date, filters)

    async def _get_admin_dashboard_data(self, start_date: datetime, end_date: datetime, filters: dict) -> dict:
        """Get admin dashboard data with filters."""
        # Build base conditions
        date_conditions = and_(
            Order.created_at >= start_date,
            Order.created_at <= end_date
        )
        
        # Apply filters to queries
        filter_conditions = []
        if filters.get('order_status'):
            filter_conditions.append(Order.status == filters['order_status'])
        
        # For category and product filters, we need to join with OrderItem and Product
        needs_product_join = filters.get('category') or filters.get('product')
        
        # Total sales
        if needs_product_join:
            sales_query = select(func.sum(OrderItem.total_price)).join(
                Order, OrderItem.order_id == Order.id
            ).join(
                ProductVariant, OrderItem.variant_id == ProductVariant.id
            ).join(
                Product, ProductVariant.product_id == Product.id
            ).where(date_conditions)
            
            if filters.get('category'):
                # Join with Category table to filter by category name
                sales_query = sales_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
            if filters.get('product'):
                sales_query = sales_query.where(Product.name.ilike(f"%{filters['product']}%"))
            if filter_conditions:
                sales_query = sales_query.where(and_(*filter_conditions))
        else:
            sales_query = select(func.sum(Order.total_amount)).where(date_conditions)
            if filter_conditions:
                sales_query = sales_query.where(and_(*filter_conditions))
                
        sales_result = await self.db.execute(sales_query)
        total_sales = sales_result.scalar() or 0

        # Total orders
        orders_query = select(func.count(Order.id.distinct())).where(date_conditions)
        if filter_conditions:
            orders_query = orders_query.where(and_(*filter_conditions))
        if needs_product_join:
            orders_query = select(func.count(Order.id.distinct())).join(
                OrderItem, Order.id == OrderItem.order_id
            ).join(
                ProductVariant, OrderItem.variant_id == ProductVariant.id
            ).join(
                Product, ProductVariant.product_id == Product.id
            ).where(date_conditions)
            if filters.get('category'):
                orders_query = orders_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
            if filters.get('product'):
                orders_query = orders_query.where(Product.name.ilike(f"%{filters['product']}%"))
            if filter_conditions:
                orders_query = orders_query.where(and_(*filter_conditions))
                
        orders_result = await self.db.execute(orders_query)
        total_orders = orders_result.scalar() or 0

        # Total users (with user segment filter)
        user_date_conditions = and_(
            User.created_at >= start_date,
            User.created_at <= end_date
        )
        users_query = select(func.count(User.id)).where(user_date_conditions)
        
        # Apply user segment filter
        if filters.get('user_segment'):
            if filters['user_segment'] == 'new':
                # Users with no orders
                users_query = select(func.count(User.id)).where(
                    and_(
                        user_date_conditions,
                        ~User.id.in_(select(Order.user_id.distinct()))
                    )
                )
            elif filters['user_segment'] == 'returning':
                # Users with more than one order
                users_query = select(func.count(User.id.distinct())).join(
                    Order, User.id == Order.user_id
                ).where(user_date_conditions).group_by(User.id).having(
                    func.count(Order.id) > 1
                )
            elif filters['user_segment'] == 'vip':
                # Users with total spending > threshold (e.g., $1000)
                users_query = select(func.count(User.id.distinct())).join(
                    Order, User.id == Order.user_id
                ).where(user_date_conditions).group_by(User.id).having(
                    func.sum(Order.total_amount) > 1000
                )
                
        users_result = await self.db.execute(users_query)
        total_users = users_result.scalar() or 0

        # Total products
        products_query = select(func.count(Product.id))
        if filters.get('category'):
            products_query = products_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        products_result = await self.db.execute(products_query)
        total_products = products_result.scalar() or 0

        # Get sales trend and top products with filters
        sales_trend = await self.get_sales_trend(
            user_id="", 
            user_role="Admin", 
            days=(end_date - start_date).days,
            filters=filters,
            start_date=start_date,
            end_date=end_date
        )
        top_products = await self.get_top_products(
            user_id="", 
            user_role="Admin", 
            limit=5,
            filters=filters
        )

        # Order status distribution
        order_status_query = select(
            Order.status,
            func.count(Order.id).label("count")
        ).where(date_conditions)
        
        if needs_product_join:
            order_status_query = select(
                Order.status,
                func.count(Order.id.distinct()).label("count")
            ).join(
                OrderItem, Order.id == OrderItem.order_id
            ).join(
                ProductVariant, OrderItem.variant_id == ProductVariant.id
            ).join(
                Product, ProductVariant.product_id == Product.id
            ).where(date_conditions)
            if filters.get('category'):
                order_status_query = order_status_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
            if filters.get('product'):
                order_status_query = order_status_query.where(Product.name.ilike(f"%{filters['product']}%"))
        
        if filter_conditions:
            order_status_query = order_status_query.where(and_(*filter_conditions))
            
        order_status_query = order_status_query.group_by(Order.status)
        order_status_result = await self.db.execute(order_status_query)
        order_status_distribution = {
            row.status: row.count for row in order_status_result.all()}

        # User growth
        user_growth_query = select(
            func.date(User.created_at).label("date"),
            func.count(User.id).label("new_users")
        ).where(user_date_conditions).group_by(
            func.date(User.created_at)
        ).order_by(
            func.date(User.created_at)
        )
        user_growth_result = await self.db.execute(user_growth_query)
        user_growth = [{
            "date": row.date.strftime("%Y-%m-%d") if hasattr(row.date, 'strftime') else str(row.date),
            "new_users": row.new_users
        } for row in user_growth_result.all()]

        # Conversion rate (simplified: unique purchasing users / total users)
        purchasing_users_query = select(func.count(Order.user_id.distinct())).where(date_conditions)
        if filter_conditions:
            purchasing_users_query = purchasing_users_query.where(and_(*filter_conditions))
        purchasing_users_result = await self.db.execute(purchasing_users_query)
        unique_purchasing_users = purchasing_users_result.scalar() or 0

        conversion_rate = (unique_purchasing_users /
                           total_users * 100) if total_users > 0 else 0.0

        return {
            "total_sales": float(total_sales),
            "total_orders": total_orders,
            "total_users": total_users,
            "total_products": total_products,
            "conversion_rate": round(conversion_rate, 2),
            "average_order_value": float(total_sales / total_orders) if total_orders > 0 else 0,
            "top_products": top_products,
            "sales_trend": sales_trend,
            "order_status_distribution": order_status_distribution,
            "user_growth": user_growth
        }

    async def _get_supplier_dashboard_data(self, user_id: str, start_date: datetime, end_date: datetime, filters: dict) -> dict:
        """Get supplier dashboard data with filters."""
        # Get supplier's products
        products_query = select(Product).where(Product.supplier_id == user_id)
        if filters.get('category'):
            products_query = products_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        products_result = await self.db.execute(products_query)
        products = products_result.scalars().all()

        # Calculate total sales and orders for the supplier
        order_item_alias = aliased(OrderItem)
        product_variant_alias = aliased(ProductVariant)

        date_conditions = and_(
            Order.created_at >= start_date,
            Order.created_at <= end_date
        )

        supplier_orders_query = select(Order).join(
            order_item_alias, Order.id == order_item_alias.order_id
        ).join(
            product_variant_alias, order_item_alias.variant_id == product_variant_alias.id
        ).join(
            Product, product_variant_alias.product_id == Product.id
        ).where(
            and_(
                Product.supplier_id == user_id,
                date_conditions
            )
        )
        
        if filters.get('category'):
            supplier_orders_query = supplier_orders_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        if filters.get('order_status'):
            supplier_orders_query = supplier_orders_query.where(Order.status == filters['order_status'])
            
        supplier_orders_query = supplier_orders_query.distinct()

        supplier_orders_result = await self.db.execute(supplier_orders_query)
        supplier_orders = supplier_orders_result.scalars().all()

        total_sales = sum(order.total_amount for order in supplier_orders)
        total_orders = len(supplier_orders)

        # Calculate average rating for supplier's products
        reviews_query = select(func.avg(Review.rating)).join(
            Product, Review.product_id == Product.id
        ).where(
            Product.supplier_id == user_id
        )
        if filters.get('category'):
            reviews_query = reviews_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        average_rating_result = await self.db.execute(reviews_query)
        average_rating = average_rating_result.scalar() or 0.0

        return {
            "total_products": len(products),
            "total_sales": float(total_sales),
            "total_orders": total_orders,
            "average_rating": round(average_rating, 2),
            "products": [
                {
                    "id": str(p.id),
                    "name": p.name,
                    "rating": p.rating,
                    "review_count": p.review_count
                }
                for p in products[:5]
            ]
        }

    async def _get_customer_dashboard_data(self, user_id: str, start_date: datetime, end_date: datetime, filters: dict) -> dict:
        """Get customer dashboard data with filters."""
        # Get customer's orders
        date_conditions = and_(
            Order.user_id == user_id,
            Order.created_at >= start_date,
            Order.created_at <= end_date
        )
        orders_query = select(Order).where(date_conditions)
        
        if filters.get('order_status'):
            orders_query = orders_query.where(Order.status == filters['order_status'])
            
        orders_result = await self.db.execute(orders_query)
        orders = orders_result.scalars().all()

        total_spent = sum(order.total_amount for order in orders)

        return {
            "total_orders": len(orders),
            "total_spent": float(total_spent),
            "average_order_value": float(total_spent / len(orders)) if orders else 0,
            "recent_orders": [
                {
                    "id": str(order.id),
                    "status": order.status,
                    "total_amount": order.total_amount,
                    "created_at": order.created_at.isoformat()
                }
                for order in orders[:5]
            ]
        }

    async def get_sales_trend(self, user_id: str, user_role: str, days: int = 30, filters: dict = None, start_date: datetime = None, end_date: datetime = None) -> List[dict]:
        """Get sales trend data grouped by date with filters."""
        if filters is None:
            filters = {}
            
        # Use provided dates if available, otherwise calculate from days
        if end_date is None:
            end_date = datetime.now()
        if start_date is None:
            start_date = end_date - timedelta(days=days)

        date_conditions = and_(
            Order.created_at >= start_date,
            Order.created_at <= end_date
        )

        # Check if we need product joins for filtering
        needs_product_join = filters.get('category') or filters.get('product') or user_role == "Supplier"

        if needs_product_join:
            # Use OrderItem-based query for product filtering
            query = select(
                func.date(Order.created_at).label("date"),
                func.sum(OrderItem.total_price).label("sales"),
                func.count(Order.id.distinct()).label("orders")
            ).join(
                OrderItem, Order.id == OrderItem.order_id
            ).join(
                ProductVariant, OrderItem.variant_id == ProductVariant.id
            ).join(
                Product, ProductVariant.product_id == Product.id
            ).where(date_conditions)
            
            if user_role == "Supplier":
                query = query.where(Product.supplier_id == user_id)
            if filters.get('category'):
                query = query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
            if filters.get('product'):
                query = query.where(Product.name.ilike(f"%{filters['product']}%"))
            if filters.get('order_status'):
                query = query.where(Order.status == filters['order_status'])
                
            query = query.group_by(
                func.date(Order.created_at)
            ).order_by(
                func.date(Order.created_at)
            )
        else:
            # Use simple Order-based query
            query = select(
                func.date(Order.created_at).label("date"),
                func.sum(Order.total_amount).label("sales"),
                func.count(Order.id).label("orders")
            ).where(date_conditions)
            
            if filters.get('order_status'):
                query = query.where(Order.status == filters['order_status'])
                
            query = query.group_by(
                func.date(Order.created_at)
            ).order_by(
                func.date(Order.created_at)
            )

        result = await self.db.execute(query)
        trend_data = []
        for row in result.all():
            trend_data.append({
                "date": row.date.strftime("%Y-%m-%d") if hasattr(row.date, 'strftime') else str(row.date),
                "sales": float(row.sales or 0),
                "orders": int(row.orders or 0)
            })

        return trend_data

    async def get_top_products(self, user_id: str, user_role: str, limit: int = 10, filters: dict = None) -> List[dict]:
        """Get top performing products by revenue with filters."""
        if filters is None:
            filters = {}
            
        # Get the most recent orders to ensure fresh data
        from sqlalchemy.orm import selectinload
        
        query = select(
            Product.id,
            Product.name,
            func.sum(OrderItem.quantity).label("total_sales_quantity"),
            func.sum(OrderItem.total_price).label("total_revenue")
        ).join(
            ProductVariant, Product.id == ProductVariant.product_id
        ).join(
            OrderItem, ProductVariant.id == OrderItem.variant_id
        ).join(
            Order, OrderItem.order_id == Order.id
        ).where(
            Order.status.in_(['confirmed', 'shipped', 'delivered', 'processing'])
        )
        
        if user_role == "Supplier":
            query = query.where(Product.supplier_id == user_id)
        if filters.get('category'):
            query = query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        if filters.get('product'):
            query = query.where(Product.name.ilike(f"%{filters['product']}%"))
        if filters.get('order_status'):
            query = query.where(Order.status == filters['order_status'])
            
        query = query.group_by(
            Product.id,
            Product.name
        ).order_by(
            func.sum(OrderItem.total_price).desc()
        ).limit(limit)

        result = await self.db.execute(query)
        products_data = []
        for row in result.all():
            # Get product image from first variant
            image_query = select(ProductVariant).where(
                ProductVariant.product_id == row.id
            ).options(selectinload(ProductVariant.images)).limit(1)
            variant_result = await self.db.execute(image_query)
            variant = variant_result.scalar_one_or_none()
            
            image_url = None
            if variant and variant.images:
                primary_image = next((img for img in variant.images if img.is_primary), None)
                image_url = primary_image.url if primary_image else (variant.images[0].url if variant.images else None)
            
            products_data.append({
                "id": str(row.id),
                "name": row.name,
                "sales": int(row.total_sales_quantity or 0),
                "revenue": float(row.total_revenue or 0),
                "image_url": image_url
            })

        return products_data

    async def export_data(self, data: dict, format: str, export_type: str) -> Tuple[bytes, str, str]:
        """
        Export analytics data to CSV or Excel format.
        
        Args:
            data: The analytics data dictionary
            format: 'csv' or 'xlsx'
            export_type: Type of export (e.g., 'dashboard')
            
        Returns:
            Tuple of (file_content, content_type, filename)
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format == 'csv':
            return await self._export_csv(data, export_type, timestamp)
        elif format == 'xlsx':
            return await self._export_excel(data, export_type, timestamp)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    async def _export_csv(self, data: dict, export_type: str, timestamp: str) -> Tuple[bytes, str, str]:
        """Export data as CSV."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write summary section
        writer.writerow(['Analytics Summary'])
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Sales', f"${data.get('total_sales', 0):.2f}"])
        writer.writerow(['Total Orders', data.get('total_orders', 0)])
        writer.writerow(['Total Users', data.get('total_users', 0)])
        writer.writerow(['Total Products', data.get('total_products', 0)])
        writer.writerow(['Conversion Rate', f"{data.get('conversion_rate', 0):.2f}%"])
        writer.writerow(['Average Order Value', f"${data.get('average_order_value', 0):.2f}"])
        writer.writerow([])
        
        # Write sales trend section
        writer.writerow(['Sales Trend'])
        writer.writerow(['Date', 'Sales', 'Orders'])
        for item in data.get('sales_trend', []):
            writer.writerow([
                item.get('date', ''),
                f"${item.get('sales', 0):.2f}",
                item.get('orders', 0)
            ])
        writer.writerow([])
        
        # Write top products section
        writer.writerow(['Top Products'])
        writer.writerow(['Product Name', 'Units Sold', 'Revenue'])
        for product in data.get('top_products', []):
            writer.writerow([
                product.get('name', ''),
                product.get('sales', 0),
                f"${product.get('revenue', 0):.2f}"
            ])
        writer.writerow([])
        
        # Write order status distribution
        writer.writerow(['Order Status Distribution'])
        writer.writerow(['Status', 'Count'])
        for status, count in data.get('order_status_distribution', {}).items():
            writer.writerow([status, count])
        
        # Convert to bytes
        content = output.getvalue().encode('utf-8')
        filename = f"analytics_{export_type}_{timestamp}.csv"
        
        return content, 'text/csv', filename
    
    async def _export_excel(self, data: dict, export_type: str, timestamp: str) -> Tuple[bytes, str, str]:
        """Export data as Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Summary"
        
        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=14)
        
        # Write summary section
        row = 1
        ws.cell(row=row, column=1, value='Analytics Summary').font = title_font
        row += 1
        
        ws.cell(row=row, column=1, value='Metric').font = header_font
        ws.cell(row=row, column=2, value='Value').font = header_font
        row += 1
        
        summary_data = [
            ('Total Sales', f"${data.get('total_sales', 0):.2f}"),
            ('Total Orders', data.get('total_orders', 0)),
            ('Total Users', data.get('total_users', 0)),
            ('Total Products', data.get('total_products', 0)),
            ('Conversion Rate', f"{data.get('conversion_rate', 0):.2f}%"),
            ('Average Order Value', f"${data.get('average_order_value', 0):.2f}")
        ]
        
        for metric, value in summary_data:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 2
        
        # Write sales trend section
        ws.cell(row=row, column=1, value='Sales Trend').font = title_font
        row += 1
        
        ws.cell(row=row, column=1, value='Date').font = header_font
        ws.cell(row=row, column=2, value='Sales').font = header_font
        ws.cell(row=row, column=3, value='Orders').font = header_font
        row += 1
        
        for item in data.get('sales_trend', []):
            ws.cell(row=row, column=1, value=item.get('date', ''))
            ws.cell(row=row, column=2, value=f"${item.get('sales', 0):.2f}")
            ws.cell(row=row, column=3, value=item.get('orders', 0))
            row += 1
        
        row += 2
        
        # Write top products section
        ws.cell(row=row, column=1, value='Top Products').font = title_font
        row += 1
        
        ws.cell(row=row, column=1, value='Product Name').font = header_font
        ws.cell(row=row, column=2, value='Units Sold').font = header_font
        ws.cell(row=row, column=3, value='Revenue').font = header_font
        row += 1
        
        for product in data.get('top_products', []):
            ws.cell(row=row, column=1, value=product.get('name', ''))
            ws.cell(row=row, column=2, value=product.get('sales', 0))
            ws.cell(row=row, column=3, value=f"${product.get('revenue', 0):.2f}")
            row += 1
        
        row += 2
        
        # Write order status distribution
        ws.cell(row=row, column=1, value='Order Status Distribution').font = title_font
        row += 1
        
        ws.cell(row=row, column=1, value='Status').font = header_font
        ws.cell(row=row, column=2, value='Count').font = header_font
        row += 1
        
        for status, count in data.get('order_status_distribution', {}).items():
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        filename = f"analytics_{export_type}_{timestamp}.xlsx"
        
        return content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename

    async def get_recent_activity(self, limit: int = 100, since: datetime = None) -> List[dict]:
        """Get recent activity logs with user details and timestamps."""
        from models.activity_log import ActivityLog
        from sqlalchemy.orm import selectinload
        
        query = select(ActivityLog).options(
            selectinload(ActivityLog.user)
        ).order_by(ActivityLog.created_at.desc())
        
        if since:
            query = query.where(ActivityLog.created_at >= since)
        
        query = query.limit(limit)
        
        result = await self.db.execute(query)
        activities = result.scalars().all()
        
        return [
            {
                "id": str(activity.id),
                "user_id": str(activity.user_id) if activity.user_id else None,
                "user_name": activity.user.full_name if activity.user else "System",
                "action_type": activity.action_type,
                "description": activity.description,
                "metadata": activity.meta_data,  # Use meta_data column name
                "created_at": activity.created_at.isoformat() if activity.created_at else None
            }
            for activity in activities
        ]
